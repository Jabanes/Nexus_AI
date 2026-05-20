"""XLSX export view (read-only, generated on demand from the DB).

The primary store of structured call data is now the SQLite DB in
src/integrations/leads/db.py. This module exists solely to materialize the
DB's contents as a downloadable workbook with two sheets:
  • "Calls"       — end-client view. Business-relevant fields only.
  • "Diagnostics" — internal/dev view. Identifiers, cost, raw paths.

The pipeline no longer calls this module on every call — it now writes
directly to the DB. Only the GET /leads/{tenant_id}/xlsx endpoint invokes
this module, building a fresh xlsx from the current DB state each time.
"""

import asyncio
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ── Sheet 1: Calls (what the end client sees) ──────────────────────────
# (header label, row dict key, column width)
_CALLS_COLUMNS: List[Tuple[str, str, int]] = [
    ("Date",            "call_date",         12),
    ("Time",            "call_time",          8),
    ("Status",          "classification",    14),
    ("Customer Name",   "customer_name",     22),
    ("Phone",           "phone_e164",        16),
    ("Address",         "address",           38),
    ("Service Needed",  "service_requested", 30),
    ("Intent",          "intent",            18),
    ("Summary",         "summary",           60),
    ("Duration (s)",    "duration_s",        12),
    ("Missing Fields",  "missing_fields",    22),
    ("Transcript",      "transcript_html",   40),
    ("Recording",       "audio_mp3",         40),
]
_CALLS_CLASSIFICATION_COL = 3   # 1-indexed column number for classification
_CALLS_HYPERLINK_COLS = (12, 13)  # transcript_html, audio_mp3


# ── Sheet 2: Diagnostics (internal / dev) ──────────────────────────────
_DIAG_COLUMNS: List[Tuple[str, str, int]] = [
    ("Session ID",       "session_id",          38),
    ("Conversation ID",  "conversation_id",     38),
    ("Agent ID",         "agent_id",            38),
    ("Tenant ID",        "tenant_id",           18),
    ("Voice Provider",   "voice_provider",      14),
    ("Started (UTC)",    "call_started_at",     22),
    ("Ended (UTC)",      "call_ended_at",       22),
    ("Termination",      "termination_reason",  28),
    ("Cost (credits)",   "cost_credits",        14),
    ("Session JSON",     "session_json",        40),
    ("Row Written",      "row_written_at",      22),
]
_DIAG_HYPERLINK_COLS = (10,)  # session_json


def _style_header_row(ws: Worksheet, columns: List[Tuple[str, str, int]]) -> None:
    """Write header row, style it, set column widths."""
    ws.append([label for (label, _key, _w) in columns])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for col_idx, (_label, _key, w) in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"


def _row_from_db_call(c: Dict[str, Any]) -> Dict[str, Any]:
    """Convert a DB row dict into the flat key/value dict the writers expect."""
    # Reconstruct call_date / call_time (Eastern) from call_started_at ISO
    call_date = ""
    call_time = ""
    started = c.get("call_started_at") or ""
    if started:
        try:
            from datetime import datetime
            dt = datetime.fromisoformat(started.replace("Z", "+00:00"))
            try:
                from zoneinfo import ZoneInfo
                dt_local = dt.astimezone(ZoneInfo("America/New_York"))
            except Exception:
                dt_local = dt
            call_date = dt_local.strftime("%Y-%m-%d")
            call_time = dt_local.strftime("%H:%M")
        except Exception:
            pass
    return {
        "call_date": call_date,
        "call_time": call_time,
        "call_started_at": c.get("call_started_at") or "",
        "call_ended_at": c.get("call_ended_at") or "",
        "duration_s": c.get("duration_s"),
        "classification": c.get("classification") or "",
        "customer_name": c.get("customer_name") or "",
        "phone_e164": c.get("phone_e164") or "",
        "address": c.get("address") or "",
        "service_requested": c.get("service_requested") or "",
        "intent": c.get("intent") or "",
        "summary": c.get("summary") or "",
        "missing_fields": c.get("missing_fields") or "",
        "termination_reason": c.get("termination_reason") or "",
        "cost_credits": c.get("cost_credits"),
        "session_id": c.get("call_id") or "",
        "conversation_id": c.get("conversation_id") or "",
        "agent_id": c.get("agent_id") or "",
        "voice_provider": c.get("voice_provider") or "",
        "tenant_id": c.get("tenant_id") or "",
        "transcript_html": c.get("transcript_html_path") or "",
        "audio_mp3": c.get("audio_mp3_path") or "",
        "session_json": c.get("session_json_path") or "",
        "row_written_at": c.get("row_written_at") or "",
    }


async def build_xlsx_from_db(tenant_id: str, dest_path: Path) -> Path:
    """Generate a fresh xlsx for a tenant from the current DB state."""
    from src.integrations.leads.db import get_db
    db = get_db()
    calls = await db.list_calls(tenant_id, limit=10_000)
    dest_path.parent.mkdir(parents=True, exist_ok=True)

    def _do():
        wb = _build_workbook()
        # DB returns newest-first; reverse to oldest-first so the sheet reads chronologically
        for call in reversed(calls):
            _append_row_to_workbook(wb, _row_from_db_call(call))
        tmp = dest_path.with_suffix(".xlsx.tmp")
        wb.save(tmp)
        tmp.replace(dest_path)

    await asyncio.to_thread(_do)
    logger.info(f"[xlsx-export] {tenant_id}: {len(calls)} rows → {dest_path}")
    return dest_path


def _build_workbook() -> Workbook:
    """Build a fresh empty workbook with both sheets and styled headers."""
    wb = Workbook()
    ws_calls = wb.active
    ws_calls.title = "Calls"
    _style_header_row(ws_calls, _CALLS_COLUMNS)
    ws_diag = wb.create_sheet("Diagnostics")
    _style_header_row(ws_diag, _DIAG_COLUMNS)
    return wb


def _append_row_to_workbook(wb: Workbook, row: Dict[str, Any]) -> None:
    _append_row_to_sheet(
        wb["Calls"], _CALLS_COLUMNS, row,
        hyperlink_cols=_CALLS_HYPERLINK_COLS,
        classification_col=_CALLS_CLASSIFICATION_COL,
    )
    _append_row_to_sheet(
        wb["Diagnostics"], _DIAG_COLUMNS, row,
        hyperlink_cols=_DIAG_HYPERLINK_COLS,
    )


def _coerce(value: Any) -> Any:
    """Normalize None / Path → string; numeric pass-through."""
    if value is None:
        return ""
    if isinstance(value, Path):
        return str(value)
    return value


def _append_row_to_sheet(
    ws: Worksheet,
    columns: List[Tuple[str, str, int]],
    row: Dict[str, Any],
    hyperlink_cols: Tuple[int, ...],
    classification_col: Optional[int] = None,
) -> None:
    """Generic row writer for either sheet."""
    values = [_coerce(row.get(key)) for (_label, key, _w) in columns]
    ws.append(values)
    row_idx = ws.max_row

    # Color-code classification (Calls sheet only)
    if classification_col is not None:
        cls = row.get("classification", "")
        cell = ws.cell(row=row_idx, column=classification_col)
        if cls == "lead":
            cell.fill = PatternFill("solid", fgColor="C8E6C9")  # green
        elif cls in ("spam", "irrelevant"):
            cell.fill = PatternFill("solid", fgColor="FFCDD2")  # red
        elif cls == "needs_review":
            cell.fill = PatternFill("solid", fgColor="FFE0B2")  # amber

    # Hyperlink file-path columns
    for col_idx in hyperlink_cols:
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            cell.hyperlink = str(cell.value)
            cell.font = Font(color="0563C1", underline="single")


